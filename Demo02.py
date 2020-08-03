import openpyxl
import requests

'''
目前了解的第三方库
requests--发送http请求并得到响应结果
jsonpath--可以做关联，取出接口响应结果里的数据
openyxl--读取excel并回写数据
'''

'''
思考一下操作excel步骤有哪些
1.打开excel
2.选择一个sheet
3.编辑
4.保存

用代码实现也是一样的操作
'''


# # 1.加载工作簿，这步骤之前可以把这个excel文件拖到这个py文件的同级目录
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# # 打印返回的是它的内存地址
# print(wb)
# # 2.选择一个sheet，来操作register这个sheet里面的内容
# sheet = wb['register']
# print(sheet)
# # 3.编辑单元格
# cell = sheet.cell(row = 1, column = 1)
# # 注意了，直接打印cell只是取的那个单元格，而不是单元格的数据
# print(cell.value)
# # 改写成另一个值
# cell.value = '测试用例编号'
# # 4.保存--打开了文件不关闭就会报错或者无法写入数据
# wb.save('test_case_api.xlsx')

# 上面只是操作一个单元格，那么怎么操作多个呢
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['register']
# url = sheet.cell(row=2,column=5).value # 取url
# data = sheet.cell(row=2,column=6).value # 取请求体
# excepted = sheet.cell(row=2,column=7).value # 取出预期结果
# print(url,data,excepted)
#
# def set_request(url, data):
#     header = {
#         'X-Lemonban-Media-Type': 'lemonban.v2',
#         'Content-Type': 'application/json'
#     }
#     response = requests.post(url=url, json=data, headers=header)
#     print(response.json())
# res = set_request(url,data)
# print(res)

# # 上面也只能一下操作一条而已
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['register']
# # 进行遍历，序列2-8左闭右开。从第二行开始取，一次循环取一次url、请求体、预期
# for i in range(2,8,1):
#     url = sheet.cell(row=i,column=5).value # 取url
#     data = sheet.cell(row=i,column=6).value # 取请求体
#     excepted = sheet.cell(row=i,column=7).value # 取出预期结果
#     print(url,data,excepted)

# # max_row 取最大行数
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['register']
# # 进行遍历，序列2-8左闭右开。从第二行开始取，一次循环取一次url、请求体、预期
# for i in range(2,sheet.max_row+1,1):
#     url = sheet.cell(row=i,column=5).value # 取url
#     data = sheet.cell(row=i,column=6).value # 取请求体
#     excepted = sheet.cell(row=i,column=7).value # 取出预期结果
#     print(url,data,excepted)

# # 取出来的数据一条一条散的，需要将这些装起来，如果再想要取哪个数据的话就比较方便
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['login']
# # 进行遍历，序列2-8左闭右开。从第二行开始取，一次循环取一次url、请求体、预期
# for i in range(2,sheet.max_row+1,1):
#     empty_dict = dict(
#     url=sheet.cell(row=i, column=5).value,  # 取url
#     data = sheet.cell(row=i, column=6).value,  # 取请求体
#     excepted = sheet.cell(row=i, column=7).value  # 取出预期结果
#     )
#     print(empty_dict)

# 大袋子
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['login']
# # 定义空列表来装
# empty_list = []
# # 进行遍历，序列2-8左闭右开。从第二行开始取，一次循环取一次url、请求体、预期
# for i in range(2,sheet.max_row+1,1):
#     empty_dict = dict(
#     url=sheet.cell(row=i, column=5).value,  # 取url
#     data = sheet.cell(row=i, column=6).value,  # 取请求体
#     excepted = sheet.cell(row=i, column=7).value  # 取出预期结果
#     )
#     empty_list.append(empty_dict)
# # 这个打印不能缩进了，之前的缩进代表还在循环里面，是因为每循环一次就要打印一次小袋子，更直观，这次只要打印最后的大袋子
# print(empty_list)

# 那么在保持表格格式不变的情况下，换另外的一个表格或者sheet呢
# def read_data(filename,sheetname):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb[sheetname]
#     # 定义空列表来装
#     empty_list = []
#     # 进行遍历，序列2-8左闭右开。从第二行开始取，一次循环取一次url、请求体、预期
#     for i in range(2,sheet.max_row+1,1):
#         empty_dict = dict(
#         url=sheet.cell(row=i, column=5).value,  # 取url
#         data = sheet.cell(row=i, column=6).value,  # 取请求体
#         excepted = sheet.cell(row=i, column=7).value  # 取出预期结果
#         )
#         empty_list.append(empty_dict)
#     # 这个打印不能缩进了，之前的缩进代表还在循环里面，是因为每循环一次就要打印一次小袋子，更直观，这次只要打印最后的大袋子
#     return empty_list
# data = read_data('test_case_api.xlsx','login')
# print(data)

# # return
# def read_data(filename,sheetname):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb[sheetname]
#     # 定义空列表来装
#     empty_list = []
#     # 进行遍历，序列2-8左闭右开。从第二行开始取，一次循环取一次url、请求体、预期
#     for i in range(2,sheet.max_row+1,1):
#         empty_dict = dict(
#         url=sheet.cell(row=i, column=5).value,  # 取url
#         data = sheet.cell(row=i, column=6).value,  # 取请求体
#         excepted = sheet.cell(row=i, column=7).value  # 取出预期结果
#         )
#         empty_list.append(empty_dict)
#     # 这个打印不能缩进了，之前的缩进代表还在循环里面，是因为每循环一次就要打印一次小袋子，更直观，这次只要打印最后的大袋子
# # return作用就是将结果返回，print只是单单的打印让你看见而已
#     return empty_list
# data = read_data('test_case_api.xlsx','login')
# print(data)

# # 写入一个数据
# wb =openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['register']
# sheet.cell(row=2,column=8).value = 'pass'
# wb.save('test_case_api.xlsx')

# # 封装成函数
# # 有变化的全设为参数
# def write_data(filename,sheetname,x,y,result):
#     wb =openpyxl.load_workbook(filename)
#     sheet = wb[sheetname]
#     sheet.cell(row=x,column=y).value = result
#     wb.save(filename)

def module(filename, sheetname):
    print('='*10 + sheetname + '模块' + '='*10)
    # 读取测试用例
    def read_data():
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheetname]
        # 定义空列表来装
        empty_list = []
        # 进行遍历，序列2-8左闭右开。从第二行开始取，一次循环取一次url、请求体、预期
        for i in range(2, sheet.max_row + 1, 1):
            empty_dict = dict(
                case_id=sheet.cell(row=i, column=1).value,  # 取id
                url=sheet.cell(row=i, column=5).value,  # 取url
                data=sheet.cell(row=i, column=6).value,  # 取请求体
                excepted=sheet.cell(row=i, column=7).value  # 取出预期结果
            )
            empty_list.append(empty_dict)
        # return作用就是将结果返回，print只是单单的打印让你看见而已
        return empty_list

    # 发送请求
    def set_request(url, data):
        header = {
            'X-Lemonban-Media-Type': 'lemonban.v2',
            'Content-Type': 'application/json'
        }
        response = requests.post(url=url, json=data, headers=header)
        # print(response.json())
        res1 = (response.json())
        return res1

    # 写入数据
    def write_data(x, y, result):
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheetname]
        sheet.cell(row=x, column=y).value = result
        wb.save(filename)

    result1 = read_data()
    for i in result1:
        # 这里格外注意：result1是大袋子--列表，i是小袋子--字典，case_id、url、data、excepted才是最后的数据
        case_id = i.get('case_id')
        url = i['url']
        data = i['data']
        data = eval(data)  # eval()函数：运行被字符串包裹的python表达式。
        excepted = i['excepted']
        excepted = eval(excepted)
        excepted_msg = excepted['msg']  # 取出预期结果里的msg
        # print(case_id,url,data,excepted)
        real_result = set_request(url, data)  # 调用set_requests函数发送请求
        real_msg = real_result['msg']  # 取出实际结果
        # 断言
        print('-' * 30)
        print('预期结果为：' + excepted_msg)
        print('实际结果为：' + real_msg)
        # 判断实际结果于预期结果是否相等
        if real_msg == excepted_msg:
            # 相等就输出第几条用例通过，并将pass写入result栏
            print('第%s条用例通过' % case_id)
            result = 'pass'
        else:
            # 不相等就输出第几条用例不通过，并将file写入result栏
            print('第%s条用例不通过' % case_id)
            result = 'fail'
        # 写入数据，调用write_data函数
        write_data(case_id + 1, 8, result)


# 调用、传参
module('test_case_api.xlsx', 'register')
module('test_case_api.xlsx', 'login')
